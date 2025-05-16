import pandas as pd
import re
from datetime import datetime
import logging
import os
import shutil
from pathlib import Path
import openpyxl
import copy

class BatchSampleProcessor:
    def __init__(self):
        self.setup_logging()
        self.setup_output_dirs()
        
    def setup_logging(self):
        log_dir = os.path.join('step2_output', 'logs')
        os.makedirs(log_dir, exist_ok=True)
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        log_file = os.path.join(log_dir, f'batch_processing_{timestamp}.log')
        
        logging.basicConfig(
            level=logging.INFO,
            format='%(asctime)s - %(levelname)s - %(message)s',
            handlers=[
                logging.FileHandler(log_file),
                logging.StreamHandler()
            ]
        )
    
    def setup_output_dirs(self):
        """Create necessary output directories"""
        self.output_dir = os.path.join('step2_output', 'processed_files')
        os.makedirs(self.output_dir, exist_ok=True)
        logging.info(f"Created output directory: {self.output_dir}")

    def extract_drug_from_cell(self, cell_content, target_drug):
        """Carefully extract specific drug from cell containing multiple drugs"""
        if not isinstance(cell_content, str):
            return None
            
        # Split cell content into individual drugs (handle both single and double spaces)
        # First replace multiple spaces with single space
        normalized = ' '.join(cell_content.replace('\n', ' ').split())
        # Then split by single space
        drugs = [d.strip() for d in normalized.split(' ') if d.strip()]
        
        # Look for exact match (case-insensitive)
        for drug in drugs:
            if drug.lower() == target_drug.lower():
                return drug
        return None

    def parse_cpic_content(self, content):
        """Parse CPIC content and return dictionary of drug recommendations"""
        logging.info("Starting to parse CPIC content...")
        recommendations = {}
        current_drug = None
        
        lines = content.split('\n')
        i = 0
        while i < len(lines):
            line = lines[i].strip()
            if not line:
                i += 1
                continue
                
            # Extract drug name when it appears alone followed by (opens in new window)
            if i < len(lines) - 1 and lines[i+1].strip() == "(opens in new window)" and not any(skip in line.lower() for skip in 
                ['learn more about', 'read full', 'see the', 'cc by-sa', 'pharmgkb', 'see the full table']):
                potential_drug = line.lower()
                if potential_drug:
                    current_drug = potential_drug
                    if current_drug not in recommendations:
                        recommendations[current_drug] = set()
                    logging.info(f"Found drug section: {current_drug}")
                i += 2  # Skip the "(opens in new window)" line
                continue
                
            # Check for CPIC recommendation start
            if current_drug and 'CPIC recommended clinical action for' in line and current_drug in line.lower():
                logging.info(f"Found CPIC recommendation line for {current_drug}: {line}")
                
                # Look ahead for Dosing Info or Alternate Drug recommendations
                j = i + 1
                while j < len(lines):
                    next_line = lines[j].strip()
                    if not next_line:
                        j += 1
                        continue
                        
                    # Stop if we hit DPWG section or next drug
                    if 'DPWG' in next_line or \
                       (j < len(lines) - 1 and lines[j+1].strip() == "(opens in new window)" and \
                        not any(skip in next_line.lower() for skip in ['learn more about', 'read full', 'see the', 'cc by-sa', 'pharmgkb'])):
                        break
                        
                    # Only look for Dosing Info or Alternate Drug
                    if 'Dosing Info' in next_line:
                        recommendations[current_drug].add('dosing_info')
                        logging.info(f"Added dosing_info recommendation for {current_drug}")
                    elif 'Alternate Drug' in next_line:
                        recommendations[current_drug].add('alternate_drug')
                        logging.info(f"Added alternate_drug recommendation for {current_drug}")
                        
                    j += 1
                    
                i = j - 1  # Resume from where we left off
                
            i += 1
        
        # Remove drugs with no actionable recommendations
        drugs_to_remove = []
        for drug in recommendations:
            if not recommendations[drug]:
                drugs_to_remove.append(drug)
                logging.info(f"Removing {drug} as it has no actionable recommendations")
        
        for drug in drugs_to_remove:
            del recommendations[drug]
        
        # Convert recommendations to list format
        final_recommendations = []
        for drug, actions in recommendations.items():
            for action in actions:
                final_recommendations.append({
                    'drug': drug,
                    'action': action
                })
        
        # Log summary of recommendations
        logging.info(f"Found {len(final_recommendations)} CPIC recommendations:")
        for rec in final_recommendations:
            logging.info(f"- {rec['drug']}: {rec['action']}")
            
        if not final_recommendations:
            logging.warning("No recommendations found in the content")
            
        return final_recommendations

    def process_excel_file(self, excel_path, recommendations):
        """Update Excel file based on recommendations"""
        try:
            # Create new output path
            filename = os.path.basename(excel_path)
            output_path = os.path.join(self.output_dir, filename)
            
            # First copy the original file to the new location
            shutil.copy2(excel_path, output_path)
            logging.info(f"Copied original file to: {output_path}")
            print(f"\nProcessing Excel file: {filename}")
            
            wb = openpyxl.load_workbook(output_path)
            ws = wb.active
            changes_made = False
            drugs_moved = []
            
            # Unmerge all cells in columns C, E, and F
            for col in ['C', 'E', 'F']:
                ranges_to_unmerge = []
                for mergedCell in ws.merged_cells.ranges:
                    if col in mergedCell.coord:
                        ranges_to_unmerge.append(str(mergedCell))
                
                for range_to_unmerge in ranges_to_unmerge:
                    # Store the style of the first cell in merged range
                    first_cell = ws[range_to_unmerge.split(':')[0]]
                    style_copy = {
                        'font': copy.copy(first_cell.font),
                        'fill': copy.copy(first_cell.fill),
                        'border': copy.copy(first_cell.border),
                        'alignment': copy.copy(first_cell.alignment),
                        'number_format': first_cell.number_format,
                        'protection': copy.copy(first_cell.protection)
                    }
                    ws.unmerge_cells(range_to_unmerge)
                    # Apply stored style to all cells in the unmerged range
                    for cell in ws[range_to_unmerge]:
                        for c in cell:
                            c.font = style_copy['font']
                            c.fill = style_copy['fill']
                            c.border = style_copy['border']
                            c.alignment = style_copy['alignment']
                            c.number_format = style_copy['number_format']
                            c.protection = style_copy['protection']
                    logging.info(f"Unmerged cells in range: {range_to_unmerge}")
            
            # Print all drugs in column C first
            print("\nDrugs found in Column C (rows 8-104):")
            for row in range(8, 105):
                cell = ws[f'C{row}']
                cell_value = cell.value
                if cell_value:
                    print(f"Row {row}: {cell_value}")
            
            # Group recommendations by drug and action
            drug_actions = {}
            for rec in recommendations:
                drug = rec['drug']
                action = rec['action']
                if drug not in drug_actions:
                    drug_actions[drug] = set()
                drug_actions[drug].add(action)
            
            # Process rows 8-104 only
            for row in range(8, 105):
                cell_c = ws[f'C{row}']
                if not cell_c.value:
                    continue
                    
                # Store cell style attributes before any modifications
                style_copy = {
                    'font': copy.copy(cell_c.font),
                    'fill': copy.copy(cell_c.fill),
                    'border': copy.copy(cell_c.border),
                    'alignment': copy.copy(cell_c.alignment),
                    'number_format': cell_c.number_format,
                    'protection': copy.copy(cell_c.protection)
                }
                
                cell_value = cell_c.value
                logging.info(f"Checking row {row}, Column C content: {cell_value}")
                print(f"Checking row {row}: {cell_value}")
                
                # Normalize and split cell value into individual drugs
                normalized = ' '.join(str(cell_value).replace('\n', ' ').split())
                drugs_in_cell = [d.strip() for d in normalized.split(' ') if d.strip()]
                
                # Keep track of drugs to remove from column C
                drugs_to_remove = []
                drugs_to_keep = []
                
                # First pass: identify which drugs have recommendations
                for drug_in_cell in drugs_in_cell:
                    found_match = False
                    for rec_drug in drug_actions:
                        if rec_drug.lower() == drug_in_cell.lower():
                            drugs_to_remove.append(drug_in_cell)
                            found_match = True
                            break
                    if not found_match:
                        drugs_to_keep.append(drug_in_cell)
                
                # Second pass: process drugs with recommendations
                for drug_in_cell in drugs_to_remove:
                    for rec_drug, actions in drug_actions.items():
                        if rec_drug.lower() == drug_in_cell.lower():
                            # Handle dosing info recommendation
                            if 'dosing_info' in actions:
                                cell_e = ws[f'E{row}']
                                current_e = cell_e.value
                                new_e = f"{current_e}\n{drug_in_cell}" if current_e else drug_in_cell
                                cell_e.value = new_e
                                # Apply stored style
                                cell_e.font = style_copy['font']
                                cell_e.fill = style_copy['fill']
                                cell_e.border = style_copy['border']
                                cell_e.alignment = style_copy['alignment']
                                cell_e.number_format = style_copy['number_format']
                                cell_e.protection = style_copy['protection']
                                logging.info(f"Moved '{drug_in_cell}' to column E (Dosing Info)")
                                print(f"Moved '{drug_in_cell}' to column E (Dosing Info)")
                                drugs_moved.append(f"{drug_in_cell} -> Column E")
                                changes_made = True
                            
                            # Handle alternate drug recommendation
                            if 'alternate_drug' in actions:
                                cell_f = ws[f'F{row}']
                                current_f = cell_f.value
                                new_f = f"{current_f}\n{drug_in_cell}" if current_f else drug_in_cell
                                cell_f.value = new_f
                                # Apply stored style
                                cell_f.font = style_copy['font']
                                cell_f.fill = style_copy['fill']
                                cell_f.border = style_copy['border']
                                cell_f.alignment = style_copy['alignment']
                                cell_f.number_format = style_copy['number_format']
                                cell_f.protection = style_copy['protection']
                                logging.info(f"Moved '{drug_in_cell}' to column F (Alternate Drug)")
                                print(f"Moved '{drug_in_cell}' to column F (Alternate Drug)")
                                drugs_moved.append(f"{drug_in_cell} -> Column F")
                                changes_made = True
                
                # Update column C with only the drugs that should stay
                if drugs_to_keep or drugs_to_remove:
                    cell_c.value = ' '.join(drugs_to_keep) if drugs_to_keep else None
                    # Apply stored style back to column C
                    cell_c.font = style_copy['font']
                    cell_c.fill = style_copy['fill']
                    cell_c.border = style_copy['border']
                    cell_c.alignment = style_copy['alignment']
                    cell_c.number_format = style_copy['number_format']
                    cell_c.protection = style_copy['protection']
                    logging.info(f"Removed {', '.join(drugs_to_remove)} from column C. Remaining drugs: {', '.join(drugs_to_keep) if drugs_to_keep else 'None'}")
                    print(f"Removed {', '.join(drugs_to_remove)} from column C. Remaining: {', '.join(drugs_to_keep) if drugs_to_keep else 'None'}")
            
            # Save changes
            try:
                wb.save(output_path)
                if changes_made:
                    logging.info(f"\nSummary of changes for {filename}:")
                    print(f"\nSummary of changes for {filename}:")
                    for move in drugs_moved:
                        logging.info(f"- {move}")
                        print(f"- {move}")
                    logging.info(f"Updated Excel file saved to: {output_path}")
                    print(f"Updated Excel file saved to: {output_path}")
                else:
                    logging.warning(f"No changes were made to {filename}")
                    print(f"No changes were made to {filename}")
            except Exception as e:
                logging.error(f"Error saving file: {str(e)}")
                print(f"Error saving file: {str(e)}")
            
        except Exception as e:
            logging.error(f"Error processing Excel file {excel_path}: {str(e)}")
            print(f"Error processing Excel file: {str(e)}")

    def record_changes_in_aggregate(self, sample_id, recommendations, changes_made):
        """Record changes in Drug Wise Aggregate Samples.xlsx"""
        try:
            aggregate_path = "/home/ronit-gandotra/Desktop/sadaf work 2/Drug Wise Aggregate Samples.xlsx"
            wb = openpyxl.load_workbook(aggregate_path)
            ws = wb.active
            
            # Find the next available column
            last_col = 1
            for cell in ws[1]:
                if cell.value:
                    last_col = cell.column
            next_col = last_col + 1
            
            # Write sample ID in the header
            ws.cell(row=1, column=next_col, value=sample_id)
            
            # Create a map of drug changes and their associated genes
            drug_changes = {}
            for rec in recommendations:
                drug = rec['drug']
                action = rec['action']
                # Extract gene info from the content file
                gene_info = self.extract_gene_info(drug)
                
                if action == 'dosing_info':
                    change = "Dosage Change"
                elif action == 'alternate_drug':
                    change = "Consider Alternate"
                    
                if drug not in drug_changes:
                    drug_changes[drug] = {'change': change, 'genes': gene_info}
                else:
                    # If drug already has a change, append the new change
                    drug_changes[drug]['change'] = f"{drug_changes[drug]['change']} & {change}"
            
            # Create a mapping of lowercase drug names to their exact names in column B
            drug_name_map = {}
            for row in range(2, ws.max_row + 1):
                cell_value = ws.cell(row=row, column=2).value
                if cell_value:
                    drug_name_map[cell_value.strip().lower()] = cell_value.strip()
            
            # Process each drug in column B
            for row in range(2, ws.max_row + 1):
                drug_cell = ws.cell(row=row, column=2)
                if not drug_cell.value:
                    continue
                    
                drug_name = drug_cell.value.strip()
                drug_name_lower = drug_name.lower()
                
                # Get the change and genes for this drug
                found_change = False
                for rec_drug, changes in drug_changes.items():
                    if rec_drug.lower() in drug_name_lower or drug_name_lower in rec_drug.lower():
                        change = changes['change']
                        genes = changes['genes']
                        value = f"{change} ({genes})" if genes else change
                        found_change = True
                        break
                
                if not found_change:
                    value = "Standard Precautions"
                
                # Write the value to the next column
                cell = ws.cell(row=row, column=next_col)
                cell.value = value
                
                # Copy style from the previous column
                prev_cell = ws.cell(row=row, column=next_col-1)
                cell.font = copy.copy(prev_cell.font)
                cell.fill = copy.copy(prev_cell.fill)
                cell.border = copy.copy(prev_cell.border)
                cell.alignment = copy.copy(prev_cell.alignment)
            
            # Save the workbook
            wb.save(aggregate_path)
            logging.info(f"Recorded changes for sample {sample_id} in Drug Wise Aggregate Samples.xlsx")
            print(f"\nUpdated Drug Wise Aggregate Samples.xlsx with changes for {sample_id}")
            
        except Exception as e:
            logging.error(f"Error recording changes in aggregate file: {str(e)}")
            print(f"Error recording changes in aggregate file: {str(e)}")
            
    def extract_gene_info(self, drug):
        """Extract gene information for a drug from the content file"""
        try:
            # Find the content file for current sample
            content_dir = os.path.join('step2_input', 'scraped_content')
            content_files = [f for f in os.listdir(content_dir) 
                           if f.startswith(f"{self.current_sample}_processed_content")]
            
            if not content_files:
                logging.error(f"No content file found for sample {self.current_sample}")
                return ""
                
            content_file = os.path.join(content_dir, content_files[0])
            
            # Read the content file
            with open(content_file, 'r') as f:
                content = f.read()
            
            # Find the drug section and extract gene info
            lines = content.split('\n')
            found_drug = False
            genes = []
            
            for i, line in enumerate(lines):
                if drug in line.lower() and "(opens in new window)" in lines[i+1]:
                    found_drug = True
                    continue
                    
                if found_drug and "CPIC recommended clinical action for" in line and drug in line.lower():
                    # Extract genes from the recommendation line
                    # Example: "CPIC recommended clinical action for fluvastatin and CYP2C9 *1/*29 and SLCO1B1 *1/*37"
                    gene_matches = re.findall(r'([A-Z0-9]+)\s+\*\d+/\*\d+', line)
                    genes.extend(gene_matches)
                    break
                    
                if found_drug and "(opens in new window)" in line:  # Next drug section
                    break
            
            return ", ".join(genes) if genes else ""
            
        except Exception as e:
            logging.error(f"Error extracting gene info for {drug}: {str(e)}")
            return ""

    def get_all_samples(self):
        """Get list of all samples from step1 output directory"""
        step1_dir = "step1_phenotype_genotype_output/processed_files"
        scraped_dir = "step2_input/scraped_content"
        
        # Get all Excel files from step1
        excel_files = [f[:-5] for f in os.listdir(step1_dir) if f.endswith('.xlsx')]
        
        # Get all scraped content files
        content_files = []
        for f in os.listdir(scraped_dir):
            if f.endswith('.txt') and '_processed_content_' in f:
                sample_id = f.split('_processed_content_')[0]
                content_files.append(sample_id)
        
        # Only process samples that have both Excel and content files
        valid_samples = list(set(excel_files) & set(content_files))
        logging.info(f"Found {len(valid_samples)} samples to process")
        return valid_samples

    def process_all_samples(self):
        """Process all samples from step1 output"""
        samples = self.get_all_samples()
        total_samples = len(samples)
        processed = 0
        failed = 0
        
        print(f"\nStarting batch processing of {total_samples} samples")
        print("="*50)
        
        for sample_id in samples:
            try:
                print(f"\nProcessing sample {processed+1}/{total_samples}: {sample_id}")
                self.current_sample = sample_id
                
                # Find content file
                content_dir = "step2_input/scraped_content"
                content_files = [f for f in os.listdir(content_dir) 
                               if f.startswith(f"{sample_id}_processed_content")]
                
                if not content_files:
                    logging.error(f"No content file found for sample {sample_id}")
                    failed += 1
                    continue
                    
                content_file = os.path.join(content_dir, content_files[0])
                
                # Read and process content
                with open(content_file, 'r') as f:
                    content = f.read()
                
                recommendations = self.parse_cpic_content(content)
                
                if recommendations:
                    # Process Excel file
                    excel_path = f"step1_phenotype_genotype_output/processed_files/{sample_id}.xlsx"
                    self.process_excel_file(excel_path, recommendations)
                    
                    # Record changes in aggregate file
                    self.record_changes_in_aggregate(sample_id, recommendations, True)
                    processed += 1
                else:
                    logging.warning(f"No recommendations found for sample {sample_id}")
                    failed += 1
                    
            except Exception as e:
                logging.error(f"Error processing sample {sample_id}: {str(e)}")
                print(f"Error processing sample {sample_id}: {str(e)}")
                failed += 1
                continue
        
        print("\n" + "="*50)
        print(f"Processing completed:")
        print(f"Total samples: {total_samples}")
        print(f"Successfully processed: {processed}")
        print(f"Failed: {failed}")
        print("="*50)

def main():
    processor = BatchSampleProcessor()
    processor.process_all_samples()

if __name__ == "__main__":
    main() 