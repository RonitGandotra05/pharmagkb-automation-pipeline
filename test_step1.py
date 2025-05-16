import pandas as pd
import openpyxl
import logging
from datetime import datetime
from pathlib import Path

def setup_logging():
    log_dir = Path('test_output/logs')
    log_dir.mkdir(parents=True, exist_ok=True)
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    log_file = log_dir / f'test_step1_{timestamp}.log'
    
    logging.basicConfig(
        level=logging.INFO,
        format='%(asctime)s - %(levelname)s - %(message)s',
        handlers=[
            logging.FileHandler(log_file),
            logging.StreamHandler()
        ]
    )

def test_single_sample():
    try:
        sample_id = "N1DL5740"
        logging.info(f"\nTesting gene matching for sample: {sample_id}")
        
        # Read CSV file
        csv_file = Path(f'Individual output_sample_results_JAPAN/{sample_id}.csv')
        if not csv_file.exists():
            logging.error(f"CSV file not found: {csv_file}")
            return
            
        # Read template
        template_file = Path('sample_result_structure.xlsx')
        if not template_file.exists():
            logging.error(f"Template file not found: {template_file}")
            return
            
        # Create output directory
        output_dir = Path('test_output')
        output_dir.mkdir(exist_ok=True)
        
        # Read CSV data
        df = pd.read_csv(csv_file)
        
        # Load template
        wb = openpyxl.load_workbook(template_file)
        ws = wb.active
        
        # Set header cells
        ws['B2'] = "N/A"
        ws['B3'] = "N/A"
        ws['B4'] = "N/A"
        ws['B5'] = sample_id
        
        logging.info("\nHeader cells updated:")
        logging.info(f"B2: {ws['B2'].value}")
        logging.info(f"B3: {ws['B3'].value}")
        logging.info(f"B4: {ws['B4'].value}")
        logging.info(f"B5: {ws['B5'].value}")
        
        # Create gene mapping for comparison
        csv_genes = {row['Gene Name']: {'genotype': row['Genotype'], 'phenotype': row['Phenotype']} for _, row in df.iterrows()}
        
        # Process rows 109-131
        logging.info("\nProcessing genes row by row:")
        for row in range(109, 132):
            gene = ws.cell(row=row, column=1).value
            if gene:
                gene = gene.strip()
                logging.info(f"\nRow {row}:")
                logging.info(f"Gene in template: {gene}")
                
                # Get current values in Excel
                current_genotype = ws.cell(row=row, column=2).value
                current_phenotype = ws.cell(row=row, column=3).value
                logging.info(f"Current values in Excel:")
                logging.info(f"  Genotype: {current_genotype}")
                logging.info(f"  Phenotype: {current_phenotype}")
                
                if gene in csv_genes:
                    csv_data = csv_genes[gene]
                    logging.info(f"Found gene in CSV:")
                    logging.info(f"  CSV Genotype: {csv_data['genotype']}")
                    logging.info(f"  CSV Phenotype: {csv_data['phenotype']}")
                    
                    # Update values
                    ws.cell(row=row, column=2).value = csv_data['genotype']
                    ws.cell(row=row, column=3).value = csv_data['phenotype']
                    
                    # Verify update
                    new_genotype = ws.cell(row=row, column=2).value
                    new_phenotype = ws.cell(row=row, column=3).value
                    logging.info(f"After update:")
                    logging.info(f"  New Genotype: {new_genotype}")
                    logging.info(f"  New Phenotype: {new_phenotype}")
                    
                    # Check if update was successful
                    if new_genotype != csv_data['genotype'] or new_phenotype != csv_data['phenotype']:
                        logging.error(f"Mismatch after update for {gene}:")
                        logging.error(f"  Expected Genotype: {csv_data['genotype']}, Got: {new_genotype}")
                        logging.error(f"  Expected Phenotype: {csv_data['phenotype']}, Got: {new_phenotype}")
                else:
                    logging.warning(f"Gene {gene} not found in CSV")
        
        # Save the updated file
        output_file = output_dir / f'{sample_id}_test.xlsx'
        wb.save(output_file)
        logging.info(f"\nSaved test output to: {output_file}")
        print(f"\nSaved test output to: {output_file}")
        
        # Final verification
        wb = openpyxl.load_workbook(output_file)
        ws = wb.active
        logging.info("\nFinal verification of saved file:")
        for row in range(109, 132):
            gene = ws.cell(row=row, column=1).value
            if gene and gene.strip() in csv_genes:
                gene = gene.strip()
                final_genotype = ws.cell(row=row, column=2).value
                final_phenotype = ws.cell(row=row, column=3).value
                csv_data = csv_genes[gene]
                if final_genotype != csv_data['genotype'] or final_phenotype != csv_data['phenotype']:
                    logging.error(f"Final verification failed for {gene}:")
                    logging.error(f"  Expected Genotype: {csv_data['genotype']}, Got: {final_genotype}")
                    logging.error(f"  Expected Phenotype: {csv_data['phenotype']}, Got: {final_phenotype}")
                else:
                    logging.info(f"Verified {gene}: Values match CSV data")
        
    except Exception as e:
        logging.error(f"Error in test_single_sample: {str(e)}")
        print(f"Error: {str(e)}")

if __name__ == "__main__":
    setup_logging()
    test_single_sample() 