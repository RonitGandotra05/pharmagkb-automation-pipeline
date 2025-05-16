import pandas as pd
import openpyxl
import logging
from datetime import datetime
from pathlib import Path

class SampleProcessor:
    def __init__(self):
        self.setup_logging()
        self.setup_paths()
        
    def setup_logging(self):
        log_dir = Path('step1_phenotype_genotype_output/logs')
        log_dir.mkdir(parents=True, exist_ok=True)
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        log_file = log_dir / f'step1_processing_{timestamp}.log'
        
        logging.basicConfig(
            level=logging.INFO,
            format='%(asctime)s - %(levelname)s - %(message)s',
            handlers=[
                logging.FileHandler(log_file),
                logging.StreamHandler()
            ]
        )
        
    def setup_paths(self):
        self.main_list_path = Path('Main Samplelist(Sheet1).csv')
        self.japan_samples_dir = Path('Individual output_sample_results_JAPAN')
        self.template_path = Path('sample_result_structure.xlsx')
        self.output_dir = Path('step1_phenotype_genotype_output/processed_files')
        self.output_dir.mkdir(parents=True, exist_ok=True)
        
    def process_samples(self):
        try:
            logging.info("Starting Step 1: Copy Phenotype and Genotype processing")
            
            # Read main sample list
            encodings = ['utf-8', 'latin1', 'cp1252', 'iso-8859-1']
            df_main = None
            
            for encoding in encodings:
                try:
                    df_main = pd.read_csv(self.main_list_path, encoding=encoding)
                    logging.info(f"Successfully read main list with {encoding} encoding")
                    break
                except UnicodeDecodeError:
                    continue
                    
            if df_main is None:
                logging.error("Failed to read main list with any encoding")
                return
                
            # Process Japan samples (rows 52-63, inclusive)
            japan_samples = df_main.iloc[50:63]  # 50 for row 52 (1-based), 63 for row 63 (inclusive)
            total_samples = len(japan_samples)
            processed_samples = 0
            missing_samples = []
            
            logging.info(f"Found {total_samples} Japan samples to process (rows 52-63, inclusive)")
            
            for idx, row in japan_samples.iterrows():
                # Use column index 1 (Column B) for Sample ID
                sample_id = str(row.iloc[1]).strip()
                actual_row = idx + 2  # Convert 0-based index to Excel row number
                logging.info(f"\nProcessing sample: {sample_id} (Row {actual_row})")  # Show actual Excel row number
                
                # Find corresponding CSV file
                csv_file = self.japan_samples_dir / f"{sample_id}.csv"
                if not csv_file.exists():
                    logging.error(f"CSV file not found for sample {sample_id}")
                    missing_samples.append(sample_id)
                    continue
                    
                try:
                    # Read CSV data
                    df = pd.read_csv(csv_file)
                    logging.info(f"Successfully read CSV file for {sample_id}")
                    
                    # Load template
                    wb = openpyxl.load_workbook(self.template_path)
                    ws = wb.active
                    
                    # Set header cells
                    ws['B2'] = "N/A"
                    ws['B3'] = "N/A"
                    ws['B4'] = "N/A"
                    ws['B5'] = sample_id
                    
                    logging.info("Updated header cells")
                    
                    # Create gene mapping
                    csv_genes = {row['Gene Name']: {'genotype': row['Genotype'], 'phenotype': row['Phenotype']} 
                               for _, row in df.iterrows()}
                    
                    # Process rows 109-131
                    logging.info(f"Processing genes for {sample_id}")
                    for row in range(109, 132):
                        gene = ws.cell(row=row, column=1).value
                        if gene:
                            gene = gene.strip()
                            if gene in csv_genes:
                                csv_data = csv_genes[gene]
                                # Update values
                                ws.cell(row=row, column=2).value = csv_data['genotype']
                                ws.cell(row=row, column=3).value = csv_data['phenotype']
                                logging.info(f"Updated {gene}: Genotype={csv_data['genotype']}, Phenotype={csv_data['phenotype']}")
                    
                    # Save the updated file
                    output_file = self.output_dir / f'{sample_id}.xlsx'
                    wb.save(output_file)
                    logging.info(f"Saved output file: {output_file}")
                    
                    # Verify the saved file
                    wb = openpyxl.load_workbook(output_file)
                    ws = wb.active
                    verification_failed = False
                    
                    for row in range(109, 132):
                        gene = ws.cell(row=row, column=1).value
                        if gene and gene.strip() in csv_genes:
                            gene = gene.strip()
                            final_genotype = ws.cell(row=row, column=2).value
                            final_phenotype = ws.cell(row=row, column=3).value
                            csv_data = csv_genes[gene]
                            
                            if final_genotype != csv_data['genotype'] or final_phenotype != csv_data['phenotype']:
                                logging.error(f"Verification failed for {gene} in {sample_id}:")
                                logging.error(f"  Expected Genotype: {csv_data['genotype']}, Got: {final_genotype}")
                                logging.error(f"  Expected Phenotype: {csv_data['phenotype']}, Got: {final_phenotype}")
                                verification_failed = True
                    
                    if not verification_failed:
                        logging.info(f"Successfully verified all genes for {sample_id}")
                        processed_samples += 1
                    
                except Exception as e:
                    logging.error(f"Error processing sample {sample_id}: {str(e)}")
                    missing_samples.append(sample_id)
                
                logging.info(f"Processed {processed_samples}/{total_samples} samples")
                
            # Generate processing report
            report_path = Path('step1_phenotype_genotype_output/processing_report.txt')
            with open(report_path, 'w', encoding='utf-8') as f:
                f.write(f"Processing Report\n")
                f.write(f"================\n")
                f.write(f"Total samples: {total_samples}\n")
                f.write(f"Successfully processed: {processed_samples}\n")
                f.write(f"Missing/Failed samples: {len(missing_samples)}\n\n")
                
                if missing_samples:
                    f.write("Missing/Failed Samples:\n")
                    for sample in missing_samples:
                        f.write(f"- {sample}\n")
                        
            logging.info("Step 1 processing completed")
            logging.info(f"Processing report saved to: {report_path}")
            
        except Exception as e:
            logging.error(f"Error in process_samples: {str(e)}")

def main():
    processor = SampleProcessor()
    processor.process_samples()

if __name__ == "__main__":
    main() 